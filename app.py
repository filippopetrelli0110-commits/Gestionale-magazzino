import csv
import io
import os
import sqlite3
from datetime import datetime
from functools import wraps
from decimal import Decimal, ROUND_HALF_UP

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:  # psycopg2 serve solo quando usi Supabase/PostgreSQL
    psycopg2 = None
    RealDictCursor = None

from flask import Flask, render_template, request, redirect, url_for, flash, session, g, send_file, Response
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.environ.get("DATABASE_PATH", os.path.join(APP_DIR, "warehouse.db"))
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = DATABASE_URL.lower().startswith(("postgresql://", "postgres://"))

if DATABASE_URL.startswith("postgres://"):
    # Alcuni servizi usano postgres://, ma postgresql:// è più compatibile.
    DATABASE_URL = "postgresql://" + DATABASE_URL[len("postgres://"):]

if psycopg2 is not None:
    DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError, psycopg2.IntegrityError)
else:
    DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError,)

app = Flask(__name__, template_folder=".", static_folder=".", static_url_path="/static")
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cambia-questa-chiave-in-produzione")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if os.environ.get("SESSION_COOKIE_SECURE", "0") == "1":
    app.config["SESSION_COOKIE_SECURE"] = True


def ensure_database_directory():
    database_dir = os.path.dirname(DATABASE_PATH)
    if database_dir:
        os.makedirs(database_dir, exist_ok=True)


def money(value):
    value = Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return value


class QueryResult:
    def __init__(self, rows=None):
        self.rows = rows or []
        self.index = 0

    def fetchone(self):
        if self.index >= len(self.rows):
            return None
        row = self.rows[self.index]
        self.index += 1
        return row

    def fetchall(self):
        if self.index == 0:
            return self.rows
        rows = self.rows[self.index:]
        self.index = len(self.rows)
        return rows


class PostgresConnection:
    def __init__(self, database_url):
        if psycopg2 is None:
            raise RuntimeError("psycopg2-binary non è installato. Esegui: pip install -r requirements.txt")
        self.database_url = self._with_sslmode(database_url)
        self.conn = psycopg2.connect(self.database_url)

    @staticmethod
    def _with_sslmode(database_url):
        # Supabase richiede normalmente SSL. Se la stringa lo contiene già, non lo duplichiamo.
        if "sslmode=" in database_url.lower():
            return database_url
        sep = "&" if "?" in database_url else "?"
        return f"{database_url}{sep}sslmode=require"

    def execute(self, query, params=()):
        params = tuple(params or ())
        pg_query = query.replace("?", "%s")
        try:
            with self.conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(pg_query, params)
                if cur.description:
                    return QueryResult([dict(row) for row in cur.fetchall()])
                return QueryResult([])
        except Exception:
            self.conn.rollback()
            raise

    def executemany(self, query, seq_of_params):
        pg_query = query.replace("?", "%s")
        try:
            with self.conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.executemany(pg_query, seq_of_params)
            return QueryResult([])
        except Exception:
            self.conn.rollback()
            raise

    def executescript(self, script):
        try:
            with self.conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(script)
            return QueryResult([])
        except Exception:
            self.conn.rollback()
            raise

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def get_db():
    if "db" not in g:
        if USE_POSTGRES:
            g.db = PostgresConnection(DATABASE_URL)
        else:
            ensure_database_directory()
            g.db = sqlite3.connect(DATABASE_PATH, timeout=30)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON")
            g.db.execute("PRAGMA busy_timeout = 30000")
            g.db.execute("PRAGMA journal_mode = WAL")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    if USE_POSTGRES:
        db = PostgresConnection(DATABASE_URL)
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin', 'operatore')),
                created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                product_code TEXT NOT NULL,
                product_name TEXT NOT NULL,
                description TEXT,
                size TEXT NOT NULL,
                purchase_price NUMERIC(12,2) NOT NULL CHECK(purchase_price >= 0),
                sale_price NUMERIC(12,2) NOT NULL CHECK(sale_price >= 0),
                min_stock INTEGER NOT NULL DEFAULT 0 CHECK(min_stock >= 0),
                created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(product_code, size)
            );

            CREATE TABLE IF NOT EXISTS stock_movements (
                id SERIAL PRIMARY KEY,
                movement_type TEXT NOT NULL CHECK(movement_type IN ('CARICO', 'SCARICO')),
                product_id INTEGER NOT NULL REFERENCES products(id) ON DELETE RESTRICT,
                quantity INTEGER NOT NULL CHECK(quantity > 0),
                recipient TEXT,
                operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                notes TEXT,
                document_number TEXT NOT NULL UNIQUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            """
        )
    else:
        ensure_database_directory()
        db = sqlite3.connect(DATABASE_PATH, timeout=30)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        db.execute("PRAGMA busy_timeout = 30000")
        db.execute("PRAGMA journal_mode = WAL")
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin', 'operatore')),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_code TEXT NOT NULL,
                product_name TEXT NOT NULL,
                description TEXT,
                size TEXT NOT NULL,
                purchase_price REAL NOT NULL CHECK(purchase_price >= 0),
                sale_price REAL NOT NULL CHECK(sale_price >= 0),
                min_stock INTEGER NOT NULL DEFAULT 0 CHECK(min_stock >= 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(product_code, size)
            );

            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                movement_type TEXT NOT NULL CHECK(movement_type IN ('CARICO', 'SCARICO')),
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL CHECK(quantity > 0),
                recipient TEXT,
                operator_id INTEGER NOT NULL,
                notes TEXT,
                document_number TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE RESTRICT,
                FOREIGN KEY(operator_id) REFERENCES users(id) ON DELETE RESTRICT
            );
            """
        )

    # Migrazioni leggere: servono se Supabase contiene tabelle create da una versione precedente.
    # CREATE TABLE IF NOT EXISTS non aggiorna automaticamente le colonne mancanti.
    if USE_POSTGRES:
        db.executescript(
            """
            ALTER TABLE products ADD COLUMN IF NOT EXISTS min_stock INTEGER NOT NULL DEFAULT 0;
            ALTER TABLE products ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP;
            """
        )
    else:
        product_cols = [r["name"] for r in db.execute("PRAGMA table_info(products)").fetchall()]
        if "min_stock" not in product_cols:
            db.execute("ALTER TABLE products ADD COLUMN min_stock INTEGER NOT NULL DEFAULT 0")
        if "updated_at" not in product_cols:
            db.execute("ALTER TABLE products ADD COLUMN updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP")

    # Utenti demo, creati solo se il database è vuoto.
    users_count = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
    if users_count == 0:
        db.execute(
            "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
            ("Amministratore", "admin@magazzino.it", generate_password_hash("admin123"), "admin"),
        )
        db.execute(
            "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
            ("Operatore", "operatore@magazzino.it", generate_password_hash("operatore123"), "operatore"),
        )

    products_count = db.execute("SELECT COUNT(*) AS c FROM products").fetchone()["c"]
    if products_count == 0:
        sample_products = [
            ("TSH001", "T-shirt bianca", "T-shirt cotone girocollo", "S", 8.00, 19.90, 5),
            ("TSH001", "T-shirt bianca", "T-shirt cotone girocollo", "M", 8.00, 19.90, 5),
            ("TSH001", "T-shirt bianca", "T-shirt cotone girocollo", "L", 8.00, 19.90, 5),
            ("FEL002", "Felpa nera", "Felpa cappuccio unisex", "M", 18.00, 39.90, 3),
        ]
        db.executemany(
            """
            INSERT INTO products
            (product_code, product_name, description, size, purchase_price, sale_price, min_stock)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            sample_products,
        )
    db.commit()
    db.close()

def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


@app.before_request
def load_logged_user():
    g.user = current_user()


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        return view(**kwargs)
    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        if g.user["role"] != "admin":
            flash("Accesso riservato all'amministratore.", "danger")
            return redirect(url_for("dashboard"))
        return view(**kwargs)
    return wrapped_view


def generate_document_number(movement_type):
    prefix = "CAR" if movement_type == "CARICO" else "SCA"
    today = datetime.now().strftime("%Y%m%d")
    db = get_db()
    count = db.execute(
        """
        SELECT COUNT(*) AS c
        FROM stock_movements
        WHERE movement_type = ? AND document_number LIKE ?
        """,
        (movement_type, f"{prefix}-{today}-%"),
    ).fetchone()["c"]
    return f"{prefix}-{today}-{count + 1:04d}"


def stock_for_product(product_id):
    row = get_db().execute(
        """
        SELECT
            COALESCE(SUM(CASE WHEN movement_type = 'CARICO' THEN quantity ELSE 0 END), 0) -
            COALESCE(SUM(CASE WHEN movement_type = 'SCARICO' THEN quantity ELSE 0 END), 0) AS stock
        FROM stock_movements
        WHERE product_id = ?
        """,
        (product_id,),
    ).fetchone()
    return int(row["stock"] or 0)


def inventory_query(where_clause="", params=()):
    return get_db().execute(
        f"""
        SELECT
            p.*,
            COALESCE(SUM(CASE WHEN sm.movement_type = 'CARICO' THEN sm.quantity ELSE 0 END), 0) -
            COALESCE(SUM(CASE WHEN sm.movement_type = 'SCARICO' THEN sm.quantity ELSE 0 END), 0) AS stock
        FROM products p
        LEFT JOIN stock_movements sm ON sm.product_id = p.id
        {where_clause}
        GROUP BY p.id
        ORDER BY p.product_name ASC, p.size ASC
        """,
        params,
    ).fetchall()


def find_product(identifier, size):
    return get_db().execute(
        """
        SELECT * FROM products
        WHERE size = ? AND (
            LOWER(product_code) = LOWER(?) OR
            LOWER(product_name) = LOWER(?)
        )
        LIMIT 1
        """,
        (size.strip(), identifier.strip(), identifier.strip()),
    ).fetchone()


def normalize_excel_header(value):
    text = str(value or "").strip().lower()
    replacements = {
        "à": "a", "è": "e", "é": "e", "ì": "i", "ò": "o", "ù": "u",
        " ": "_", "-": "_", ".": "_", "/": "_",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def first_present(row, names, default=""):
    for name in names:
        if name in row and row[name] is not None:
            value = row[name]
            if isinstance(value, str):
                value = value.strip()
            if value != "":
                return value
    return default


def to_int_quantity(value):
    if value is None or value == "":
        raise ValueError("quantita mancante")
    try:
        qty = int(float(str(value).replace(",", ".")))
    except Exception:
        raise ValueError("quantita non valida")
    if qty <= 0:
        raise ValueError("quantita non positiva")
    return qty


def to_float_price(value, default=0):
    if value is None or value == "":
        return float(default)
    text = str(value).strip().replace("€", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    return float(text)


def create_or_update_product_from_import(code, name, description, size, purchase_price, sale_price, min_stock=0):
    db = get_db()
    existing = db.execute(
        "SELECT * FROM products WHERE LOWER(product_code) = LOWER(?) AND LOWER(size) = LOWER(?) LIMIT 1",
        (code, size),
    ).fetchone()
    if existing:
        # Se il prodotto esiste, aggiorniamo solo dati descrittivi/prezzi se arrivano dal file.
        db.execute(
            """
            UPDATE products
            SET product_name = COALESCE(NULLIF(?, ''), product_name),
                description = COALESCE(NULLIF(?, ''), description),
                purchase_price = ?,
                sale_price = ?,
                min_stock = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (name, description, purchase_price, sale_price, int(min_stock or 0), existing["id"]),
        )
        return db.execute("SELECT * FROM products WHERE id = ?", (existing["id"],)).fetchone(), False

    db.execute(
        """
        INSERT INTO products
        (product_code, product_name, description, size, purchase_price, sale_price, min_stock)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (code, name, description, size, purchase_price, sale_price, int(min_stock or 0)),
    )
    return db.execute(
        "SELECT * FROM products WHERE LOWER(product_code) = LOWER(?) AND LOWER(size) = LOWER(?) LIMIT 1",
        (code, size),
    ).fetchone(), True


def inventory_rows():
    rows = []
    for p in inventory_query():
        stock = int(p["stock"] or 0)
        purchase_value = money(stock * p["purchase_price"])
        sale_value = money(stock * p["sale_price"])
        margin = sale_value - purchase_value
        rows.append({
            "code": p["product_code"],
            "name": p["product_name"],
            "description": p["description"] or "",
            "size": p["size"],
            "stock": stock,
            "purchase_price": money(p["purchase_price"]),
            "sale_price": money(p["sale_price"]),
            "purchase_value": purchase_value,
            "sale_value": sale_value,
            "margin": margin,
            "min_stock": p["min_stock"],
        })
    return rows


@app.route("/__status")
def status():
    try:
        db = get_db()
        users = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
        products = db.execute("SELECT COUNT(*) AS c FROM products").fetchone()["c"]
        return {
            "ok": True,
            "database": "postgres/supabase" if USE_POSTGRES else "sqlite",
            "users": int(users),
            "products": int(products),
        }
    except Exception as exc:
        return {
            "ok": False,
            "error": type(exc).__name__,
            "message": str(exc),
        }, 500


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.user:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_db().execute("SELECT * FROM users WHERE LOWER(email) = ?", (email,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            flash(f"Accesso effettuato. Benvenuto, {user['name']}.", "success")
            return redirect(url_for("dashboard"))
        flash("Email o password non corretti.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logout effettuato.", "success")
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    products = inventory_query()
    total_variants = len(products)
    total_units = sum(int(p["stock"] or 0) for p in products)
    total_purchase_value = sum(money(int(p["stock"] or 0) * p["purchase_price"]) for p in products)
    total_sale_value = sum(money(int(p["stock"] or 0) * p["sale_price"]) for p in products)
    low_stock = [p for p in products if int(p["stock"] or 0) <= int(p["min_stock"] or 0)]
    last_loads = get_db().execute(
        """
        SELECT sm.*, p.product_code, p.product_name, p.size, u.name AS operator_name
        FROM stock_movements sm
        JOIN products p ON p.id = sm.product_id
        JOIN users u ON u.id = sm.operator_id
        WHERE sm.movement_type = 'CARICO'
        ORDER BY sm.created_at DESC, sm.id DESC LIMIT 5
        """
    ).fetchall()
    last_unloads = get_db().execute(
        """
        SELECT sm.*, p.product_code, p.product_name, p.size, u.name AS operator_name
        FROM stock_movements sm
        JOIN products p ON p.id = sm.product_id
        JOIN users u ON u.id = sm.operator_id
        WHERE sm.movement_type = 'SCARICO'
        ORDER BY sm.created_at DESC, sm.id DESC LIMIT 5
        """
    ).fetchall()
    return render_template(
        "dashboard.html",
        total_variants=total_variants,
        total_units=total_units,
        total_purchase_value=money(total_purchase_value),
        total_sale_value=money(total_sale_value),
        low_stock=low_stock,
        last_loads=last_loads,
        last_unloads=last_unloads,
    )


@app.route("/products", methods=["GET", "POST"])
@login_required
def products():
    if request.method == "POST":
        if g.user["role"] != "admin":
            flash("Solo l'amministratore può creare articoli.", "danger")
            return redirect(url_for("products"))
        data = request.form
        try:
            get_db().execute(
                """
                INSERT INTO products
                (product_code, product_name, description, size, purchase_price, sale_price, min_stock)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data.get("product_code", "").strip().upper(),
                    data.get("product_name", "").strip(),
                    data.get("description", "").strip(),
                    data.get("size", "").strip().upper(),
                    float(data.get("purchase_price", 0)),
                    float(data.get("sale_price", 0)),
                    int(data.get("min_stock", 0) or 0),
                ),
            )
            get_db().commit()
            flash("Articolo creato correttamente.", "success")
        except DB_INTEGRITY_ERRORS:
            flash("Esiste già un articolo con lo stesso codice prodotto e la stessa taglia.", "danger")
        except ValueError:
            flash("Controlla prezzi e soglia minima: devono essere numeri validi.", "danger")
        return redirect(url_for("products"))

    q = request.args.get("q", "").strip()
    if q:
        rows = inventory_query(
            "WHERE LOWER(p.product_code) LIKE LOWER(?) OR LOWER(p.product_name) LIKE LOWER(?) OR LOWER(p.size) LIKE LOWER(?)",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        )
    else:
        rows = inventory_query()
    return render_template("products.html", products=rows, q=q)


@app.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_product(product_id):
    product = get_db().execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if not product:
        flash("Articolo non trovato.", "danger")
        return redirect(url_for("products"))
    if request.method == "POST":
        data = request.form
        try:
            get_db().execute(
                """
                UPDATE products
                SET product_code = ?, product_name = ?, description = ?, size = ?,
                    purchase_price = ?, sale_price = ?, min_stock = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    data.get("product_code", "").strip().upper(),
                    data.get("product_name", "").strip(),
                    data.get("description", "").strip(),
                    data.get("size", "").strip().upper(),
                    float(data.get("purchase_price", 0)),
                    float(data.get("sale_price", 0)),
                    int(data.get("min_stock", 0) or 0),
                    product_id,
                ),
            )
            get_db().commit()
            flash("Articolo aggiornato.", "success")
            return redirect(url_for("products"))
        except DB_INTEGRITY_ERRORS:
            flash("Codice prodotto + taglia già utilizzati.", "danger")
        except ValueError:
            flash("Controlla prezzi e soglia minima.", "danger")
    return render_template("edit_product.html", product=product)


@app.route("/products/<int:product_id>/delete", methods=["POST"])
@admin_required
def delete_product(product_id):
    movements = get_db().execute("SELECT COUNT(*) AS c FROM stock_movements WHERE product_id = ?", (product_id,)).fetchone()["c"]
    if movements > 0:
        flash("Non puoi eliminare un articolo con movimenti registrati. Puoi modificarlo o lasciarlo nello storico.", "danger")
    else:
        get_db().execute("DELETE FROM products WHERE id = ?", (product_id,))
        get_db().commit()
        flash("Articolo eliminato.", "success")
    return redirect(url_for("products"))


@app.route("/load", methods=["GET", "POST"])
@login_required
def load_stock():
    if request.method == "POST":
        identifier = request.form.get("identifier", "")
        size = request.form.get("size", "")
        notes = request.form.get("notes", "").strip()
        try:
            quantity = int(request.form.get("quantity", "0"))
            if quantity <= 0:
                raise ValueError
        except ValueError:
            flash("La quantità deve essere un numero intero positivo.", "danger")
            return redirect(url_for("load_stock"))

        product = find_product(identifier, size)
        if not product:
            flash("Articolo non trovato: controlla codice/nome e taglia.", "danger")
            return redirect(url_for("load_stock"))

        document_number = generate_document_number("CARICO")
        get_db().execute(
            """
            INSERT INTO stock_movements
            (movement_type, product_id, quantity, recipient, operator_id, notes, document_number)
            VALUES ('CARICO', ?, ?, NULL, ?, ?, ?)
            """,
            (product["id"], quantity, g.user["id"], notes, document_number),
        )
        get_db().commit()
        flash(f"Carico salvato. Documento {document_number}.", "success")
        return redirect(url_for("movements", movement_type="CARICO"))
    return render_template("load.html")


@app.route("/load/template.xlsx")
@login_required
def load_template_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Carico"
    headers = [
        "codice_prodotto", "nome_prodotto", "descrizione", "taglia", "quantita",
        "prezzo_acquisto", "prezzo_vendita", "soglia_minima", "note"
    ]
    ws.append(headers)
    ws.append(["TSH001", "T-shirt bianca", "T-shirt cotone girocollo", "S", 10, 8.00, 19.90, 5, "primo carico"])
    ws.append(["TSH001", "T-shirt bianca", "T-shirt cotone girocollo", "M", 15, 8.00, 19.90, 5, "primo carico"])
    ws.append(["FEL002", "Felpa nera", "Felpa cappuccio unisex", "L", 8, 18.00, 39.90, 3, ""])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 24, 34, 10, 12, 18, 18, 14, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '€ #,##0.00'
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="modello_carico_magazzino.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/load/import", methods=["GET", "POST"])
@login_required
def import_load_excel():
    report = None
    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        if not uploaded or uploaded.filename == "":
            flash("Carica un file Excel .xlsx.", "danger")
            return redirect(url_for("import_load_excel"))
        if not uploaded.filename.lower().endswith(".xlsx"):
            flash("Formato non valido. Usa un file .xlsx.", "danger")
            return redirect(url_for("import_load_excel"))

        try:
            wb = load_workbook(uploaded, data_only=True)
            ws = wb.active
        except Exception:
            flash("File Excel non leggibile. Controlla che sia un .xlsx valido.", "danger")
            return redirect(url_for("import_load_excel"))

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            flash("Il file Excel è vuoto.", "danger")
            return redirect(url_for("import_load_excel"))
        headers = [normalize_excel_header(h) for h in header_row]
        rows_to_import = []
        errors = []

        for excel_row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            try:
                code = str(first_present(row, ["codice_prodotto", "codice", "codice_articolo", "sku", "code"])).strip().upper()
                name = str(first_present(row, ["nome_prodotto", "nome", "prodotto", "articolo", "product_name"])).strip()
                description = str(first_present(row, ["descrizione", "descrizione_prodotto", "description"], "")).strip()
                size = str(first_present(row, ["taglia", "size"], "")).strip().upper()
                qty = to_int_quantity(first_present(row, ["quantita", "quantita_caricata", "pezzi", "qta", "quantity"]))
                purchase_price = to_float_price(first_present(row, ["prezzo_acquisto", "costo", "purchase_price"], 0), 0)
                sale_price = to_float_price(first_present(row, ["prezzo_vendita", "vendita", "sale_price"], 0), 0)
                min_stock = int(float(str(first_present(row, ["soglia_minima", "min_stock", "scorta_minima"], 0)).replace(",", ".")))
                notes = str(first_present(row, ["note", "notes"], "")).strip()
                if not code:
                    raise ValueError("codice_prodotto mancante")
                if not name:
                    raise ValueError("nome_prodotto mancante")
                if not size:
                    raise ValueError("taglia mancante")
                if purchase_price < 0 or sale_price < 0:
                    raise ValueError("prezzi negativi")
                rows_to_import.append({
                    "row": excel_row_number,
                    "code": code,
                    "name": name,
                    "description": description,
                    "size": size,
                    "quantity": qty,
                    "purchase_price": purchase_price,
                    "sale_price": sale_price,
                    "min_stock": max(min_stock, 0),
                    "notes": notes,
                })
            except Exception as exc:
                errors.append(f"Riga {excel_row_number}: {exc}")

        if errors:
            report = {"imported": 0, "created": 0, "updated": 0, "errors": errors, "rows": []}
            return render_template("import_load.html", report=report)

        imported = 0
        created = 0
        updated = 0
        imported_rows = []
        try:
            for item in rows_to_import:
                product, was_created = create_or_update_product_from_import(
                    item["code"], item["name"], item["description"], item["size"],
                    item["purchase_price"], item["sale_price"], item["min_stock"]
                )
                document_number = generate_document_number("CARICO")
                get_db().execute(
                    """
                    INSERT INTO stock_movements
                    (movement_type, product_id, quantity, recipient, operator_id, notes, document_number)
                    VALUES ('CARICO', ?, ?, NULL, ?, ?, ?)
                    """,
                    (product["id"], item["quantity"], g.user["id"], item["notes"], document_number),
                )
                imported += 1
                if was_created:
                    created += 1
                else:
                    updated += 1
                imported_rows.append({**item, "document_number": document_number, "created": was_created})
            get_db().commit()
            report = {"imported": imported, "created": created, "updated": updated, "errors": [], "rows": imported_rows}
            flash(f"Importazione completata: {imported} righe caricate, {created} articoli creati, {updated} articoli aggiornati.", "success")
        except Exception as exc:
            get_db().rollback()
            report = {"imported": 0, "created": 0, "updated": 0, "errors": [str(exc)], "rows": []}
            flash("Importazione annullata: nessun carico è stato salvato.", "danger")

    return render_template("import_load.html", report=report)


@app.route("/unload", methods=["GET", "POST"])
@login_required
def unload_stock():
    if request.method == "POST":
        identifier = request.form.get("identifier", "")
        size = request.form.get("size", "")
        recipient = request.form.get("recipient", "").strip()
        notes = request.form.get("notes", "").strip()
        if not recipient:
            flash("Inserisci il nome del destinatario.", "danger")
            return redirect(url_for("unload_stock"))
        try:
            quantity = int(request.form.get("quantity", "0"))
            if quantity <= 0:
                raise ValueError
        except ValueError:
            flash("La quantità deve essere un numero intero positivo.", "danger")
            return redirect(url_for("unload_stock"))

        product = find_product(identifier, size)
        if not product:
            flash("Articolo non trovato: controlla codice/nome e taglia.", "danger")
            return redirect(url_for("unload_stock"))

        available = stock_for_product(product["id"])
        if quantity > available:
            flash(f"Giacenza insufficiente. Disponibili: {available} pezzi.", "danger")
            return redirect(url_for("unload_stock"))

        document_number = generate_document_number("SCARICO")
        get_db().execute(
            """
            INSERT INTO stock_movements
            (movement_type, product_id, quantity, recipient, operator_id, notes, document_number)
            VALUES ('SCARICO', ?, ?, ?, ?, ?, ?)
            """,
            (product["id"], quantity, recipient, g.user["id"], notes, document_number),
        )
        get_db().commit()
        movement_id = get_db().execute(
            "SELECT id FROM stock_movements WHERE document_number = ?",
            (document_number,),
        ).fetchone()["id"]
        flash(f"Scarico salvato. Documento {document_number}.", "success")
        return redirect(url_for("print_movement", movement_id=movement_id))
    return render_template("unload.html")


@app.route("/movements")
@login_required
def movements():
    filters = {
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "code": request.args.get("code", "").strip(),
        "name": request.args.get("name", "").strip(),
        "size": request.args.get("size", "").strip(),
        "operator": request.args.get("operator", "").strip(),
        "recipient": request.args.get("recipient", "").strip(),
        "movement_type": request.args.get("movement_type", "").strip(),
    }
    clauses = []
    params = []
    if filters["date_from"]:
        clauses.append("DATE(sm.created_at) >= DATE(?)")
        params.append(filters["date_from"])
    if filters["date_to"]:
        clauses.append("DATE(sm.created_at) <= DATE(?)")
        params.append(filters["date_to"])
    if filters["code"]:
        clauses.append("LOWER(p.product_code) LIKE LOWER(?)")
        params.append(f"%{filters['code']}%")
    if filters["name"]:
        clauses.append("LOWER(p.product_name) LIKE LOWER(?)")
        params.append(f"%{filters['name']}%")
    if filters["size"]:
        clauses.append("LOWER(p.size) LIKE LOWER(?)")
        params.append(f"%{filters['size']}%")
    if filters["operator"]:
        clauses.append("LOWER(u.name) LIKE LOWER(?)")
        params.append(f"%{filters['operator']}%")
    if filters["recipient"]:
        clauses.append("LOWER(COALESCE(sm.recipient, '')) LIKE LOWER(?)")
        params.append(f"%{filters['recipient']}%")
    if filters["movement_type"] in ("CARICO", "SCARICO"):
        clauses.append("sm.movement_type = ?")
        params.append(filters["movement_type"])
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    rows = get_db().execute(
        f"""
        SELECT sm.*, p.product_code, p.product_name, p.description, p.size,
               u.name AS operator_name
        FROM stock_movements sm
        JOIN products p ON p.id = sm.product_id
        JOIN users u ON u.id = sm.operator_id
        {where}
        ORDER BY sm.created_at DESC, sm.id DESC
        """,
        params,
    ).fetchall()
    return render_template("movements.html", movements=rows, filters=filters)


@app.route("/movements/<int:movement_id>/print")
@login_required
def print_movement(movement_id):
    movement = get_db().execute(
        """
        SELECT sm.*, p.product_code, p.product_name, p.description, p.size,
               u.name AS operator_name
        FROM stock_movements sm
        JOIN products p ON p.id = sm.product_id
        JOIN users u ON u.id = sm.operator_id
        WHERE sm.id = ?
        """,
        (movement_id,),
    ).fetchone()
    if not movement:
        flash("Movimento non trovato.", "danger")
        return redirect(url_for("movements"))
    return render_template("print_movement.html", m=movement)


@app.route("/exports")
@login_required
def exports():
    rows = inventory_rows()
    totals = {
        "stock": sum(r["stock"] for r in rows),
        "purchase_value": sum(r["purchase_value"] for r in rows),
        "sale_value": sum(r["sale_value"] for r in rows),
        "margin": sum(r["margin"] for r in rows),
    }
    return render_template("exports.html", rows=rows, totals=totals)


@app.route("/exports/inventory.csv")
@login_required
def export_csv():
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Codice prodotto", "Nome prodotto", "Descrizione", "Taglia", "Quantità disponibile",
        "Prezzo acquisto", "Prezzo vendita", "Valore acquisto", "Valore vendita", "Margine teorico"
    ])
    for r in inventory_rows():
        writer.writerow([
            r["code"], r["name"], r["description"], r["size"], r["stock"],
            r["purchase_price"], r["sale_price"], r["purchase_value"], r["sale_value"], r["margin"]
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=giacenze_magazzino.csv"},
    )


@app.route("/exports/inventory.xlsx")
@login_required
def export_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Giacenze"
    headers = [
        "Codice prodotto", "Nome prodotto", "Descrizione", "Taglia", "Quantità disponibile",
        "Prezzo acquisto", "Prezzo vendita", "Valore acquisto", "Valore vendita", "Margine teorico"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")
    for r in inventory_rows():
        ws.append([
            r["code"], r["name"], r["description"], r["size"], r["stock"],
            float(r["purchase_price"]), float(r["sale_price"]), float(r["purchase_value"]),
            float(r["sale_value"]), float(r["margin"])
        ])
    total_row = ws.max_row + 2
    totals = {
        "stock": sum(r["stock"] for r in inventory_rows()),
        "purchase_value": sum(r["purchase_value"] for r in inventory_rows()),
        "sale_value": sum(r["sale_value"] for r in inventory_rows()),
        "margin": sum(r["margin"] for r in inventory_rows()),
    }
    ws.cell(total_row, 1, "TOTALI")
    ws.cell(total_row, 5, totals["stock"])
    ws.cell(total_row, 8, float(totals["purchase_value"]))
    ws.cell(total_row, 9, float(totals["sale_value"]))
    ws.cell(total_row, 10, float(totals["margin"]))
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = '€ #,##0.00'
    thin = Side(style="thin", color="CBD5E1")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="giacenze_magazzino.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/exports/inventory/print")
@login_required
def export_print():
    rows = inventory_rows()
    totals = {
        "stock": sum(r["stock"] for r in rows),
        "purchase_value": sum(r["purchase_value"] for r in rows),
        "sale_value": sum(r["sale_value"] for r in rows),
        "margin": sum(r["margin"] for r in rows),
    }
    return render_template("print_inventory.html", rows=rows, totals=totals)


@app.route("/users", methods=["GET", "POST"])
@admin_required
def users():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        role = request.form.get("role", "operatore")
        if role not in ("admin", "operatore"):
            role = "operatore"
        if not name or not email or not password:
            flash("Compila nome, email e password.", "danger")
            return redirect(url_for("users"))
        try:
            get_db().execute(
                "INSERT INTO users (name, email, password_hash, role) VALUES (?, ?, ?, ?)",
                (name, email, generate_password_hash(password), role),
            )
            get_db().commit()
            flash("Utente creato.", "success")
        except DB_INTEGRITY_ERRORS:
            flash("Esiste già un utente con questa email.", "danger")
        return redirect(url_for("users"))
    rows = get_db().execute("SELECT id, name, email, role, created_at FROM users ORDER BY name").fetchall()
    return render_template("users.html", users=rows)


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_user(user_id):
    if user_id == g.user["id"]:
        flash("Non puoi eliminare il tuo stesso utente.", "danger")
        return redirect(url_for("users"))
    movements = get_db().execute("SELECT COUNT(*) AS c FROM stock_movements WHERE operator_id = ?", (user_id,)).fetchone()["c"]
    if movements > 0:
        flash("Non puoi eliminare un utente collegato a movimenti già registrati.", "danger")
    else:
        get_db().execute("DELETE FROM users WHERE id = ?", (user_id,))
        get_db().commit()
        flash("Utente eliminato.", "success")
    return redirect(url_for("users"))


@app.template_filter("eur")
def eur_filter(value):
    try:
        value = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"€ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "€ 0,00"


@app.template_filter("dt")
def dt_filter(value):
    if not value:
        return ""
    try:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M")
        # SQLite restituisce stringhe; Supabase/PostgreSQL può restituire datetime.
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        try:
            return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
        except Exception:
            return value


# Inizializza il database anche quando l'app parte online con Gunicorn.
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
